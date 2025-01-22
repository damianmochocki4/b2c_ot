import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from collections import defaultdict
from datetime import datetime
import calendar
from dateutil.relativedelta import relativedelta
import locale
from copy import copy
import pandas as pd
import re
import time
import math

# Record the start time
start_time = time.time()

#set month names to be German
locale.setlocale(locale.LC_TIME, 'de_DE')

# Dictionary mapping project names to filenames
projects = {
    'Abnahme MFG': '',
    'IBN PoP': '',
    'Prüfung FNT': '',
    'Eingang Doku': ''
}

# Initialize variables
ot_file = ''
reference_date = None

if getattr(sys, 'frozen', False):
    # If running as a bundled executable (e.g., PyInstaller)
    # Workaround for PyInstaller onefile apps that extract to temp folder
    directory = os.path.dirname(os.path.abspath(sys.argv[0]))
    if sys.argv[0].endswith(".exe"):
        # Sometimes, the bundled executable is unpacked into a temporary directory
        # Use this condition to handle extraction if needed.
        pass
else:
    # If running as a normal script
    directory = os.path.dirname(os.path.abspath(__file__))

# Function to extract reference date from the filename
def extract_reference_date_from_filename(filename):
    # Search for a sequence of 8 digits in the filename
    match = re.search(r'\d{8}', filename)
    if match:
        # Extract the matched date string
        date_str = match.group()
        # Convert the date string to a datetime object
        reference_date = datetime.strptime(date_str, "%d%m%Y")
        return reference_date
    else:
        # Return None if no date is found
        return None

def browse_ot_file():
    global ot_file
    filename = filedialog.askopenfilename(title="Select OT File")
    if filename:
        ot_file = filename
        ot_file_entry.delete(0, tk.END)
        ot_file_entry.insert(0, filename)

def browse_project_file(project):
    global projects
    filename = filedialog.askopenfilename(title=f"Select {project} File")
    if filename:
        projects[project] = filename
        project_entries[project].delete(0, tk.END)
        project_entries[project].insert(0, filename)
        
def browse_bst_file():
    global bst_file_path
    filename = filedialog.askopenfilename(title="Select BST File")
    if filename:
        bst_file_path = filename
        bst_file_entry.delete(0, tk.END)
        bst_file_entry.insert(0, filename)
    else:
        bst_file_path = os.path.join(directory, "BST.xlsx")

def submit():
    global ot_file_path, project_files, ot_file, bst_file_path
    ot_file_path = ot_file_entry.get()
    project_files = {project: entry.get() for project, entry in project_entries.items() if project != 'OT File'}
    bst_file_path = bst_file_entry.get()
    
    # Check if any field is empty
    if not ot_file_path:
        error_label.config(text="Please fill in OT File path field.")
        return
    
    if not bst_file_path:
        error_label.config(text="Please fill in BST File path field.")
        return

    if not os.path.exists(ot_file_path):
        error_label.config(text="OT file does not exist.")
        return
        
    if not os.path.exists(bst_file_path):
        error_label.config(text="BST file does not existd.")
        return

    # Proceed with loading ot_file
    ot_file = load_workbook(ot_file_path)
    
    # Extract reference date from OT file name
    global reference_date
    reference_date = extract_reference_date_from_filename(os.path.basename(ot_file_path))
    root.destroy()

root = tk.Tk()
root.attributes('-topmost', True)
root.title("B2C OT Milestones Report Generator")
root.geometry("1050x200")
root.resizable(False, False)

# OT File label
tk.Label(root, text="OT File:").grid(row=0, column=0, sticky='w', padx=10)
ot_file_entry = tk.Entry(root, width=140)
ot_file_entry.grid(row=0, column=1, columnspan=4)
tk.Button(root, text="Browse", command=browse_ot_file).grid(row=0, column=5)

# Project file entries
project_entries = {}
for i, (project, _) in enumerate(projects.items(), start=1):
    tk.Label(root, text=f"{project} File:").grid(row=i, column=0, sticky='w', padx=10)
    project_entry = tk.Entry(root, width=140)
    project_entry.grid(row=i, column=1, columnspan=4)
    project_entries[project] = project_entry
    tk.Button(root, text="Browse", command=lambda proj=project: browse_project_file(proj)).grid(row=i, column=5)

# BST File
bst_file_path = os.path.join(directory, "BST.xlsx")
tk.Label(root, text="BST File:").grid(row=len(projects) + 1, column=0, sticky='w', padx=10)
bst_file_entry = tk.Entry(root, width=140)
bst_file_entry.insert(0, bst_file_path)
bst_file_entry.grid(row=len(projects) + 1, column=1, columnspan=4)
tk.Button(root, text="Browse", command=browse_bst_file).grid(row=len(projects) + 1, column=5)

# Submit button
tk.Button(root, text="Run the processing", command=submit).grid(row=len(projects) + 2, column=0, columnspan=6)

# Error label
error_label = tk.Label(root, text="", fg="red")
error_label.grid(row=len(projects) + 3, column=0, columnspan=6)

root.mainloop()

# Proceed with your script using reference_date and ot_file_path
if reference_date is None or ot_file_path == "":
    print("Reference date or OT file not specified.")
    sys.exit()

bst_file = load_workbook(bst_file_path)
bst_worksheet = bst_file[bst_file.sheetnames[0]]

changelog = ''
for project, ms in projects.items():
    
    if not ms:
        print(f"Path for '{project}' file was not provided.\n")
        changelog += f"Path for '{project}' file was not provided.\n\n"
        continue  # Skip processing this project

    #load excel file
    ms_file = load_workbook(ms)

    #create new worksheet to store Omni Tracker data in
    if not f'{reference_date.strftime("%d%m%Y")} {project}' in ms_file.sheetnames:
        ms_file.create_sheet(f'{reference_date.strftime("%d%m%Y")} {project}')
        ms_file.move_sheet(f'{reference_date.strftime("%d%m%Y")} {project}', -(len(ms_file.sheetnames)-2))
    
    #open worksheets
    ot_worksheet = ot_file[f'OT {project}']
    ms_dashboard = ms_file.worksheets[0]
    new_worksheet = ms_file.worksheets[1]
    
    #find max and min dates
    end_dates = [
        end_date.value
        for row in zip(
            ot_worksheet.iter_cols(min_col=5, max_col=5, min_row=2, max_row=ot_worksheet.max_row),
            ot_worksheet.iter_cols(min_col=7, max_col=7, min_row=2, max_row=ot_worksheet.max_row)
        )
        for project, end_date in zip(row[0], row[1])
        if 'B2C' in str(project.value) and end_date.value
    ]
    newest_date = max(end_dates)
    
    reference_month = reference_date.month
    reference_year = reference_date.year
    
    #find the most recent date (without counting the reference month)
    most_recent_date = None
    for date in end_dates:
        if date.year < reference_year or (date.year == reference_year and date.month < reference_month):
            if most_recent_date is None or date < most_recent_date: #zmieniony znak z date > most_recent_date
                most_recent_date = date 
    
    #check how many months between two dates
    def months_between_dates(start_date, end_date):
        months = (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month)
        return months
                
    if most_recent_date is None or months_between_dates(most_recent_date, reference_date) >= 2:
        most_recent_date = reference_date - relativedelta(months = 1)
    
    #range of months
    start_year, start_month = most_recent_date.year, most_recent_date.month
    end_year, end_month = newest_date.year, newest_date.month
    
    #find from which column to start from
    ms_year = None
    for i in range(1, ms_dashboard.max_column + 1):
        if ms_dashboard.cell(2, i).value:
            ms_year = ms_dashboard.cell(2, i).value
        if ms_dashboard.cell(3, i).value == calendar.month_name[start_month] and ms_year == start_year:
            ms_start_column = i
            break
    
    new_worksheet_start_column = column_index_from_string('F')
    
    #find first not hidden column
    for col_idx in range(3, ms_dashboard.max_column + 1):
        if not ms_dashboard.column_dimensions[get_column_letter(col_idx)].hidden:
            hide_col = col_idx
            break
    
    #hide not needed columns
    while hide_col != ms_start_column:
        ms_dashboard.column_dimensions[get_column_letter(hide_col)].hidden = True
        hide_col += 1
    
    #difference between starting columns in both worksheets
    start_diff = ms_start_column - new_worksheet_start_column
    
    new_worksheet['A4'] = 'Projekt'
    new_worksheet['B4'] = 'Anzahl AO'
    new_worksheet['C4'] = 'Summe Aufträge'
    new_worksheet['D4'] = 'Start (PLAN)'
    new_worksheet['E4'] = 'Ende (PLAN)'
    
    #make headers bold in the new worksheet
    for i in range(1, 6):
        new_worksheet.cell(4, i).font = Font(b=True)
        
    #copy dates from OT
    row = 5
    for i in range(2, ot_worksheet.max_row + 1):
        if ot_worksheet['E' + str(i)].value and 'B2C' in ot_worksheet['E' + str(i)].value and ot_worksheet['G' + str(i)].value:
            new_worksheet['A' + str(row)] = ot_worksheet['E' + str(i)].value
            if ot_worksheet['F' + str(i)].value:
                start_date = ot_worksheet['F' + str(i)].value.date()
                new_worksheet['D' + str(row)] = datetime.strptime(start_date.strftime('%d.%m.%Y'), '%d.%m.%Y').date()
                new_worksheet['D' + str(row)].number_format = 'DD.MM.YYYY'
            else:
                continue
            if ot_worksheet['G' + str(i)].value:
                end_date = ot_worksheet['G' + str(i)].value.date()
                if end_date.year > start_year or (end_date.year == start_year and end_date.month >= start_month):
                    new_worksheet['E' + str(row)] = datetime.strptime(end_date.strftime('%d.%m.%Y'), '%d.%m.%Y').date()
                    new_worksheet['E' + str(row)].number_format = 'DD.MM.YYYY'
                else:
                    new_worksheet.delete_rows(row)
                    row -= 1
            else:
                continue
        else:
            continue
        
        row += 1
    
    #bst_worksheet
    for i in range(5, new_worksheet.max_row + 1):
        for j in range(6, bst_worksheet.max_row + 1):
            ms_projekt = new_worksheet['A' + str(i)].value
            bst_projekt = bst_worksheet['A' + str(j)].value
            if ms_projekt and bst_projekt:
                ms_projekt = str(ms_projekt).lower().strip()
                bst_projekt = str(bst_projekt).lower().strip()
                # bst_projekt = bst_projekt.replace('neukieritzsch - b2c - cluster', 'neukieritzsch - b2c - pop-nekz1-1')
                bst_projekt = bst_projekt.replace('cluster', 'pop1')
                bst_projekt = bst_projekt.replace('pegau-elstertrebnitz', 'pegau/elstertrebnitz')
                bst_projekt = bst_projekt.replace('neukieritzsch - b2c - pop ', 'neukieritzsch - b2c - pop-nekz1-')
                if 'pop ' in ms_projekt:
                    ms_projekt = ms_projekt.replace('pop ', 'pop')
                if 'pop ' in bst_projekt:
                    bst_projekt = bst_projekt.replace('pop ', 'pop')
                    
                # ms_projekt = ms_projekt.replace('neukieritzsch - b2c - pop-nekz1-', 'neukieritzsch - b2c - pop ')
                
                if ms_projekt.endswith(bst_projekt):
                    new_worksheet['B' + str(i)] = bst_worksheet['B' + str(j)].value
                    new_worksheet['C' + str(i)] = bst_worksheet['E' + str(j)].value
    
    #load the data into a DataFrame in order to sort by end date
    df = pd.DataFrame(new_worksheet.values)
    
    #extract the header row
    header_row = df.iloc[3]
    
    #skip the header row and sort the data
    sorted_df = df.iloc[4:].sort_values(by=4, ascending=True)
    
    #reset index after sorting
    sorted_df.reset_index(drop=True, inplace=True)
    
    #insert the header row back into the sorted DataFrame
    sorted_df = pd.concat([header_row.to_frame().T, sorted_df], ignore_index=True)
    
    #overwrite the worksheet with the sorted data, preserving styling and the header row
    for r_idx, row in enumerate(dataframe_to_rows(sorted_df, index=False, header=False), 1):
        for c_idx, value in enumerate(row, 1):
            new_worksheet.cell(row=r_idx + 3, column=c_idx, value=value)
    
    month_counts = defaultdict(int)
    
    #store dates and find occurences of B2C in months
    for row in ot_worksheet.iter_rows(min_row=2, values_only=True):
        if row[4] and row[6]:
            text = str(row[4])
            date_str = str(row[6])
            
            date_str = date_str.split(' ')[0]
            date = datetime(
                int(date_str.split('-')[0]),
                int(date_str.split('-')[1]),
                int(date_str.split('-')[2]))
        
            if 'B2C' in text:
                month_counts[(date.year, date.month)] += 1
            else:
                continue
            
    #change width of start date and end date columns
    new_worksheet.column_dimensions['A'].width = 52
    new_worksheet.column_dimensions['B'].width = 14
    new_worksheet.column_dimensions['C'].width = 18
    new_worksheet.column_dimensions['D'].width = 16
    new_worksheet.column_dimensions['E'].width = 20
    
    #insert a new row and a date
    for i in range(10, ms_dashboard.max_row + 2):
        if not ms_dashboard.cell(i, 1).value:    
            ms_dashboard.insert_rows(i, 1)
            ms_dashboard.cell(i, 1).value = datetime.strptime(reference_date.strftime('%d.%m.%Y'), '%d.%m.%Y').date()
            ms_dashboard.cell(i, 1).number_format = 'DD.MM.YYYY'
            row = i
            break
    
    #hide the oldest row (it should always show the last 5 entries)
    ms_dashboard.row_dimensions[row - 5].hidden = True
    
    #create a list to store the results
    results_row = []
    months = []
    for year in range(start_year, end_year + 1):
        start_month_index = start_month if year == start_year else 1
        end_month_index = end_month if year == end_year else 12
        
        for month in range(start_month_index, end_month_index + 1):
            months.append(month)
            results_row.append(str(month_counts[(year, month)]))
    
    #insert the results in the dashboard
    for i, value in enumerate(results_row, start = 1):
        ms_dashboard.cell(row, ms_start_column + i - 1, int(value))
    
    thick_borders = Border(left=Side(style='medium'),
                         right=Side(style='medium'),
                         top=Side(style='medium'),
                         bottom=Side(style='medium')
                         )
    
    thin_borders = Border(
                        left=Side(border_style='thin'),
                        right=Side(border_style='thin'),
                        top=Side(border_style='thin'),
                        bottom=Side(border_style='thin')
                        )
    
    thick_left = Border(left=Side(style='medium'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin')
                         )
    
    thick_bottom = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='medium')
                         )
    
    thick_bottom_and_left = Border(left=Side(style='medium'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='medium')
                         )
    
    center_alignment = Alignment(horizontal='center')
    left_alignment = Alignment(horizontal='left')
    
    #change style of new row
    for i in range(1, ms_dashboard.max_column + 1):
        ms_dashboard.cell(row, i).font = copy(ms_dashboard.cell(row - 1, i).font)
        ms_dashboard.cell(row, i).border = copy(ms_dashboard.cell(row - 1, i).border)
        ms_dashboard.cell(row, i).number_format = copy(ms_dashboard.cell(row - 1, i).number_format)
        ms_dashboard.cell(row, i).protection = copy(ms_dashboard.cell(row - 1, i).protection)
        ms_dashboard.cell(row, i).alignment = copy(ms_dashboard.cell(row - 1, i).alignment)
        if i >= ms_start_column:
            ms_dashboard.column_dimensions[get_column_letter(i)].width = 12
            
        #change color of 0 and max value
        if ms_dashboard.cell(row, i).value == 0:
            ms_dashboard.cell(row, i).fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid') #green
        
        if ms_dashboard.cell(row, i).value == max([int(value) for value in results_row]):
            ms_dashboard.cell(row, i).fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') #yellow

    #insert months in the dashboard and in the new worksheet
    month_index = 0
    end_column = ms_start_column + len(months) - 1
    new_columns = []
    januarys_ms = []
    # current_year = ms_dashboard.cell(2, ms_start_column).value
    # current_year = datetime.now().year
    current_year = start_year
    for i in range(ms_start_column, end_column + 1):
        if month_index <= len(months) - 1:
            if not ms_dashboard.cell(3, i).value:
                new_columns.append(i)
            ms_dashboard.cell(3, i).value = calendar.month_name[months[month_index]]
            new_worksheet.cell(4, i - start_diff).value = calendar.month_name[months[month_index]]
            month_index += 1
            if ms_dashboard.cell(3, i).value == 'Januar':
                januarys_ms.append(i)
                if i != ms_start_column:
                    current_year += 1
                    ms_dashboard.cell(2, i).value = current_year
        else:
            break
        
    #change color of the previous months
    i = ms_start_column
    while ms_dashboard.cell(3, i).value != reference_date.strftime('%B'):
        ms_dashboard.cell(row, i).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid') #red
        i += 1

    ms_dashboard.cell(row, 2).value = f'=SUM({get_column_letter(ms_start_column)}{row}:{get_column_letter(end_column)}{row})'
    
    if len(new_columns) > 0:
        #unmerge and merge the newest year again
        if januarys_ms[-1] < end_column - len(new_columns):
            ms_dashboard.unmerge_cells(f'{get_column_letter(januarys_ms[-1])}2:{get_column_letter(end_column - len(new_columns))}2') # !! there's might be an error if there's a new year
        ms_dashboard.merge_cells(f'{get_column_letter(januarys_ms[-1])}2:{get_column_letter(end_column)}2')
        #change style of new columns
        for i in range(10, row + 1):
            for j in new_columns:
                ms_dashboard.cell(i, j).border = thin_borders
                ms_dashboard.cell(i, j).alignment = center_alignment
        
        #change style of headers    
        for i in new_columns:
            ms_dashboard.cell(3, i).border = thin_borders
            ms_dashboard.cell(3, i).alignment = center_alignment
            
        ms_dashboard.cell(2, januarys_ms[-1]).alignment = center_alignment                                  
        ms_dashboard.cell(2, januarys_ms[-1]).border = thick_borders
        
    #insert the results in the new worksheet
    for i, value in enumerate(results_row, start = 1):
        new_worksheet.cell(3, new_worksheet_start_column + i - 1, int(value))
    
    januarys_ed = []
    years_sum = []
    for i in range(6, new_worksheet.max_column + 1):
        if new_worksheet.cell(4, i).value:
            new_worksheet.column_dimensions[get_column_letter(i)].width = 12
        if not new_worksheet.cell(4, i + 1).value:
            new_worksheet_end_column = i
        if new_worksheet.cell(4, i).value == 'Dezember' and i == new_worksheet_start_column:
            years_sum.append(new_worksheet.cell(3, i).value)
        if new_worksheet.cell(4, i - 1).value == 'Dezember' and new_worksheet.cell(4, i).value == 'Januar':
            januarys_ed.append(i)
    
    # Sort the merged cell ranges by min_row
    sorted_ranges = sorted(ms_dashboard.merged_cells.ranges, key=lambda x: x.min_col)
    
    #add merged cells to the new worksheet and add the header with styles       
    merged_ranges = []
    for merged_range in sorted_ranges:
        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        if min_col >= ms_start_column and merged_range not in merged_ranges:
            merged_ranges.append(merged_range)
        if min_col < ms_start_column and max_col > ms_start_column and merged_range not in merged_ranges:
            ms_dashboard.unmerge_cells(str(merged_range))
            merged_range.min_col = ms_start_column
            ms_dashboard[get_column_letter(ms_start_column) + '2'] = ms_dashboard['A' + str(row)].value.year
            ms_dashboard[get_column_letter(ms_start_column) + '2'].alignment = center_alignment
            ms_dashboard[get_column_letter(ms_start_column) + '2'].border = thick_borders
            ms_dashboard.merge_cells(f'{get_column_letter(ms_start_column)}2:{get_column_letter(max_col)}2')
            merged_ranges.append(merged_range)

    # Filter merged_ranges to exclude ranges beyond new_worksheet_end_column
    merged_ranges = [
        merged_range for merged_range in merged_ranges 
        if merged_range.max_col <= (new_worksheet_end_column + start_diff)
    ]

    #make a sum for every year
    for merged_range in merged_ranges:
        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        new_worksheet.merge_cells(start_row=min_row, start_column=min_col - start_diff, end_row=max_row, end_column=max_col - start_diff)
        new_worksheet.cell(row=min_row, column=min_col - start_diff).value = 'Anzahl PoP pro Monat'
        new_worksheet.cell(row=min_row, column=min_col - start_diff).alignment = center_alignment
        
        values_to_sum = [new_worksheet.cell(row=min_row + 1, column=col).value for col in range(min_col - start_diff, max_col - start_diff + 1)]
        values_to_sum = [0 if value is None else int(value) for value in values_to_sum]
        
        year_sum = sum(values_to_sum)
        years_sum.append(year_sum)

    # Iterate through merged cell ranges to check if the first column is merged
    first_col_is_merged = False
    for merged_range in new_worksheet.merged_cells.ranges:
        if new_worksheet[get_column_letter(new_worksheet_start_column) + '2'].coordinate in merged_range:
            first_col_is_merged = True
            break
    
    if not first_col_is_merged:
        new_worksheet[get_column_letter(new_worksheet_start_column) + '2'] = 'Anzahl PoP pro Monat'
        new_worksheet[get_column_letter(new_worksheet_start_column) + '2'].alignment = left_alignment
    
    # Iterate through merged cell ranges to check if the last column is merged
    last_col_is_merged = False
    for merged_range in new_worksheet.merged_cells.ranges:
        if new_worksheet[get_column_letter(new_worksheet_end_column) + '2'].coordinate in merged_range:
            last_col_is_merged = True
            break
        
    if not last_col_is_merged:
        new_worksheet[get_column_letter(new_worksheet_end_column) + '2'] = 'Anzahl PoP pro Monat'
        new_worksheet[get_column_letter(new_worksheet_end_column) + '2'].alignment = left_alignment
    
    #exception for when the last column is january
    for i in range(6, new_worksheet.max_column + 1):
        if new_worksheet.cell(4, i).value == 'Januar' and i == new_worksheet_end_column:
            years_sum.append(new_worksheet.cell(3, i).value) 
        
    sum_row = 4
    years_sum_accumulated = []     
    for i in years_sum:
        sum_row += i
        years_sum_accumulated.append(sum_row)
    
    #set borders
    for row in range(4, new_worksheet.max_row + 1):
        for col in range(new_worksheet_start_column, new_worksheet.max_column + 1):
            if new_worksheet.cell(4, col).value and new_worksheet.cell(row, 5).value:
                new_worksheet.cell(row=row, column=col).border = thin_borders
                
    for i in januarys_ed:
        for row in range(4, new_worksheet.max_row + 1):
            if new_worksheet.cell(4, i).value and new_worksheet.cell(row, 5).value:
                new_worksheet.cell(row, i).border = thick_left
                
    for col in range(6, new_worksheet_end_column + 1):
        for i in years_sum_accumulated:
            if new_worksheet.cell(4, col).value and new_worksheet.cell(i, 5).value:
                new_worksheet.cell(i, col).border = thick_bottom
                if new_worksheet.cell(4, col).value == 'Januar' and col == 6:
                    continue
                elif new_worksheet.cell(4, col).value == 'Januar':
                    new_worksheet.cell(i, col).border = thick_bottom_and_left
    
    #add 2 new rows
    new_worksheet.insert_rows(3, amount=2)
    
    new_worksheet['E3'] = 'Summe Aufträge'
    new_worksheet['E4'] = 'Anzahl AO'
    new_worksheet['E5'] = 'Anzahl PoP Cluster'
    
    #insert a numbering of projects amount          
    count = 1
    column = new_worksheet_start_column
    i = 7
    auftrage_count = 0
    ao_count = 0
    while i <= new_worksheet.max_row:
        if new_worksheet.cell(6, column).value:
            if count == new_worksheet.cell(5, column).value:
                if new_worksheet.cell(i, 2).value:
                    ao_count += new_worksheet.cell(i, 2).value
                    auftrage_count += new_worksheet.cell(i, 3).value
                    new_worksheet[get_column_letter(column) + '3'] = auftrage_count
                    new_worksheet[get_column_letter(column) + '4'] = ao_count
                    auftrage_count = 0
                    ao_count = 0
                else:
                    new_worksheet[get_column_letter(column) + '3'] = auftrage_count
                    new_worksheet[get_column_letter(column) + '4'] = ao_count
                    auftrage_count = 0
                    ao_count = 0
                new_worksheet.cell(i, column).value = count
                column += 1
                count = 1
                i += 1  # Move to the next row
            elif new_worksheet.cell(5, column).value == 0 and i == 7:
                new_worksheet[get_column_letter(column) + '3'] = auftrage_count
                new_worksheet[get_column_letter(column) + '4'] = ao_count
                column += 1
            elif new_worksheet.cell(5, column).value == 0:
                new_worksheet[get_column_letter(column) + '3'] = auftrage_count
                new_worksheet[get_column_letter(column) + '4'] = ao_count
                new_worksheet.cell(i - 1, column).value = 0  # Insert 0 in the current row
                column += 1  # Move to the next column
                count = 1  # Reset count
            else:
                if new_worksheet.cell(i, 2).value:
                    ao_count += new_worksheet.cell(i, 2).value
                    auftrage_count += new_worksheet.cell(i, 3).value
                new_worksheet.cell(i, column).value = count
                count += 1
                i += 1  # Move to the next row
        else:
            # Handle the case when new_worksheet.cell(5, column).value is falsy
            i += 1  # Move to the next row
            
    #change borders and alignment
    for col in range(new_worksheet_start_column, new_worksheet_end_column + 1):
        for row in range(2, 7):
            if row == 2 or row == 6:
                new_worksheet.cell(row, col).font = Font(b=True)
            if row > 2:
                new_worksheet.cell(row, col).alignment = center_alignment
            new_worksheet.cell(row, col).border = thick_borders
    
    #change font color and background
    for row in range(7, new_worksheet.max_row + 1):
        for col in range(new_worksheet_start_column, new_worksheet.max_column + 1):
            if new_worksheet.cell(row, col).value:
                new_worksheet.cell(row=row, column=col).alignment = center_alignment
                new_worksheet.cell(row=row, column=col).font = Font(color='0C6705')
                new_worksheet.cell(row=row, column=col).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            if new_worksheet.cell(row=row, column=col).value == 0:
                new_worksheet.cell(row=row, column=col).alignment = center_alignment
    
    ao_sum_row = 0
    auftrage_sum_row = 0
    #coloring green if AO > Auftrage, red if else (for projects and month sums)            
    for row in range(7, new_worksheet.max_row + 1):
        if new_worksheet['B' + str(row)].value and new_worksheet['C' + str(row)].value:
            ao_sum_row += new_worksheet['B' + str(row)].value
            auftrage_sum_row += new_worksheet['C' + str(row)].value
            if new_worksheet['B' + str(row)].value > new_worksheet['C' + str(row)].value:
                new_worksheet['B' + str(row)].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') #green
            else:
                new_worksheet['B' + str(row)].fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid') #red
     
    ao_sum_col = 0
    auftrage_sum_col = 0
    for col in range (new_worksheet_start_column, new_worksheet.max_column + 1):
        if new_worksheet[get_column_letter(col) + '4'].value and new_worksheet[get_column_letter(col) + '3'].value:
            ao_sum_col += new_worksheet[get_column_letter(col) + '4'].value
            auftrage_sum_col += new_worksheet[get_column_letter(col) + '3'].value
            if new_worksheet[get_column_letter(col) + '4'].value > new_worksheet[get_column_letter(col) + '3'].value:
                new_worksheet[get_column_letter(col) + '4'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') #green
            else:
                new_worksheet[get_column_letter(col) + '4'].fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid') #red
    
    new_worksheet['B3'] = 'Totals'
    new_worksheet['B3'].alignment = Alignment(horizontal='center', vertical='center')
    new_worksheet.merge_cells('B3:C4')
    
    new_worksheet['B5'] = ao_sum_row
    new_worksheet['C5'] = auftrage_sum_row
    
    new_worksheet['D3'] = auftrage_sum_col
    new_worksheet['D4'] = ao_sum_col
    
    # Define the font you want to set as default
    default_font_name = 'Arial'
    default_font_size = 11
    
    # Iterate over each worksheet in the workbook
    for ws in ms_file.worksheets:
        # Apply the font to all cells in the worksheet
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                existing_font = cell.font
                # Create a new Font object preserving the existing attributes
                new_font = Font(name=default_font_name, 
                                size=default_font_size,
                                bold=existing_font.bold,
                                italic=existing_font.italic,
                                vertAlign=existing_font.vertAlign,
                                underline=existing_font.underline,
                                strike=existing_font.strike,
                                color=existing_font.color)
                cell.font = new_font
    
    # Append the current month abbreviation before the extension
    new_ms = os.path.basename(ms)[:-9] + "_" + reference_date.strftime("%B").upper()[:3] + ".xlsx"
    
    new_ms_export = os.path.join(directory, new_ms)
    
    #save the file
    ms_file.save(new_ms)
    
    print(f"'{project}' has been exported to:\n '{new_ms_export}.'\n")
    changelog += f"'{project}' has been exported to:\n '{new_ms_export}.'\n\n"

'''
OT File modification
'''

bau_worksheet = ot_file['OT Baufortschritt']

#load the data into a DataFrame
df = pd.DataFrame(bau_worksheet.values)

#extract the header row
header_row = df.iloc[0]

#skip the header row and sort the data
sorted_df = df.iloc[1:].sort_values(by=[11, 6, 4], ascending=[False, True, True])

#reset index after sorting
sorted_df.reset_index(drop=True, inplace=True)

#insert the header row back into the sorted DataFrame
sorted_df = pd.concat([header_row.to_frame().T, sorted_df], ignore_index=True)

#overwrite the worksheet with the sorted data, preserving styling and the header row
for r_idx, row in enumerate(dataframe_to_rows(sorted_df, index=False, header=False), 1):
    for c_idx, value in enumerate(row, 1):
        bau_worksheet.cell(row=r_idx, column=c_idx, value=value)

#insert ID column
bau_worksheet.insert_cols(1)
bau_worksheet.column_dimensions['A'].width = 10

bau_worksheet['A1'] = 'ID'
bau_worksheet['A1'].font = Font(b=True)
bau_worksheet['A1'].alignment = Alignment(horizontal='center')
 
#apply a filter to show rows where 'Projektphase' contains 'B2C' and delete the unnecessary row
i = 2
while i <= bau_worksheet.max_row:
    cell_value = bau_worksheet['F' + str(i)].value
    if cell_value:
        if 'Testprojekt4 B2C mit Vorlage' in cell_value:
            bau_worksheet.delete_rows(i)
        elif 'B2C' in cell_value:
            bau_worksheet.row_dimensions[i].hidden = False
            i += 1
        else:
            bau_worksheet.row_dimensions[i].hidden = True
            i += 1
    else:
        empty_row = i
        break

#calculate the average value
percentage_values = [cell.value for cell in bau_worksheet['M'][1:] if cell.value and not bau_worksheet.row_dimensions[cell.row].hidden and str(cell.value) != '0%']
average_percentage = sum(percentage_values) / len(percentage_values)

bau_worksheet['I' + str(empty_row + 1)] = 'Average:'
bau_worksheet['M' + str(empty_row + 1)] = average_percentage
bau_worksheet['M' + str(empty_row + 1)].number_format = '0%'

#iterate through each row in the ID column and assign values starting from 1 for not hidden rows
visible_row_count = 0
for row in bau_worksheet.iter_rows(min_row=2, max_row=bau_worksheet.max_row, min_col=1, max_col=1):
    cell = row[0]
    row_num = cell.row
    if not bau_worksheet.row_dimensions[row_num].hidden:
        visible_row_count += 1
        cell.value = visible_row_count
    else:
        cell.value = None
    cell.font = Font(b=True)
    cell.alignment = Alignment(horizontal='center')
    if visible_row_count == 21:
        break

#iterate through F to N columns and correct number format
reference_quarter = ((reference_date.month - 1) // 3) + 1
start_column = column_index_from_string('F')
for row_idx, row in enumerate(bau_worksheet.iter_rows(min_row=2, min_col=start_column, max_col=column_index_from_string('N')), start=2):
    if row[start_column - start_column].value:
        #highlight the dates in the current calendar quarter with yellow color
        date_value_G = row[column_index_from_string('G') - start_column].value
        if isinstance(date_value_G, datetime):
            bau_worksheet[f'G{row_idx}'].number_format = 'DD.MM.YYYY'
            quarter_G = ((date_value_G.month - 1) // 3) + 1
            if quarter_G == reference_quarter and date_value_G.year == reference_date.year:
                bau_worksheet[f'G{row_idx}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") #yellow
        
        #highlight the dates maximum 3 months in the future of the current calendar month in column H - "Ende (PLAN)" with blue color
        date_value_H = row[column_index_from_string('H') - start_column].value
        if isinstance(date_value_H, datetime):
            bau_worksheet[f'H{row_idx}'].number_format = 'DD.MM.YYYY'
            difference_in_months = (date_value_H.year - reference_date.year) * 12 + (date_value_H.month - reference_date.month)
            if difference_in_months == 0 and (date_value_H.day - reference_date.day) > 0:
                bau_worksheet[f'H{row_idx}'].fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid") #blue
            if 1 <= difference_in_months <= 3:
                bau_worksheet[f'H{row_idx}'].fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid") #blue
                
        #highlight the empty cells in column I - "Start IST" with red color
        date_value_I = row[column_index_from_string('I') - start_column].value
        if isinstance(date_value_I, datetime):
            bau_worksheet[f'I{row_idx}'].number_format = 'DD.MM.YYYY'
        if not row[column_index_from_string('I') - start_column].value:
            row[column_index_from_string('I') - start_column].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") #red
                
        #highlight the dates older than 30 days from now in column N - "Letzte Änderung" with orange color
        date_value_N = row[column_index_from_string('N') - start_column].value
        if isinstance(date_value_N, datetime):
            bau_worksheet[f'N{row_idx}'].number_format = 'DD.MM.YYYY HH:MM:SS'
            difference = reference_date - date_value_N
            if difference.days > 30:
                row[column_index_from_string('N') - start_column].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid") #orange

#insert a filter
bau_worksheet.auto_filter.ref = bau_worksheet.dimensions
        
#hide unnecessary columns & make column titles bold
columns_to_hide = [
    'Aktivität',
    'Verantwortlich',
    'Zustand',
    'Projektphase',
    'Ende (IST)',
    'Zeitaufwand PLAN (h)',
    'Zeitaufwand IST (h)'
]

for i in range(2, bau_worksheet.max_column + 1):
    if bau_worksheet[get_column_letter(i) + '1'].value in columns_to_hide:
        bau_worksheet.column_dimensions[get_column_letter(i)].hidden = True
    else:
        bau_worksheet[get_column_letter(i) + '1'].font = Font(b=True)
        
bau_worksheet.column_dimensions['F'].width = 62
bau_worksheet.column_dimensions['G'].width = 15
bau_worksheet.column_dimensions['H'].width = 15
bau_worksheet.column_dimensions['I'].width = 15
bau_worksheet.column_dimensions['M'].width = 20
bau_worksheet.column_dimensions['N'].width = 22

#create a new worksheet for the PoP Check task
if not 'PoP Check' in ot_file.sheetnames:
    ot_file.create_sheet('PoP Check')
    ot_file.move_sheet('PoP Check', -(len(ot_file.sheetnames)-1))
    
#create a new worksheet for the PoP Termine task
if not 'PoP Termine' in ot_file.sheetnames:
    ot_file.create_sheet('PoP Termine')
    ot_file.move_sheet('PoP Termine', -(len(ot_file.sheetnames)-2))

#load different worksheets for the 2nd task    
check_worksheet = ot_file['PoP Check']
termine_worksheet = ot_file['PoP Termine']
ed_worksheet = ot_file['OT Eingang Doku']
fnt_worksheet = ot_file['OT Prüfung FNT']
ab_worksheet = ot_file['OT Abnahme MFG']
ibn_worksheet = ot_file['OT IBN PoP']

#create headers for pop termine worksheet with styles
termine_worksheet['A2'] = 'Aktivität'
termine_worksheet['B2'] = 'Projekt'
termine_worksheet['C2'] = 'Ende (PLAN)'

ed_projects = []
fnt_projects = []
ab_projects = []
ibn_projects = []
all_projects = []

#fill PoP Check and PoP Termine with projects
row_check = 3
row_termine = 3
for i in range(2, ed_worksheet.max_row + 1):
    if not ed_worksheet['E' + str(i)].value:
        break
    if 'B2C' in ed_worksheet['E' + str(i)].value and ed_worksheet['G' + str(i)].value and 'Testprojekt4 B2C mit Vorlage' not in ed_worksheet['E' + str(i)].value:
        termine_worksheet['A' + str(row_termine)] = ed_worksheet['A' + str(i)].value
        termine_worksheet['B' + str(row_termine)] = ed_worksheet['E' + str(i)].value
        termine_worksheet['C' + str(row_termine)] = ed_worksheet['G' + str(i)].value
        termine_worksheet['C' + str(row_termine)].number_format = 'DD.MM.YYYY'
        row_termine += 1
        if ed_worksheet['E' + str(i)].value not in ed_projects:
            ed_projects.append(ed_worksheet['E' + str(i)].value)
        if ed_worksheet['E' + str(i)].value not in all_projects:
            all_projects.append(ed_worksheet['E' + str(i)].value)
            check_worksheet['B' + str(row_check)] = ed_worksheet['E' + str(i)].value
            row_check += 1
            
for i in range(2, fnt_worksheet.max_row + 1):
    if not fnt_worksheet['E' + str(i)].value:
        break
    if 'B2C' in fnt_worksheet['E' + str(i)].value and fnt_worksheet['G' + str(i)].value and 'Testprojekt4 B2C mit Vorlage' not in fnt_worksheet['E' + str(i)].value:
        termine_worksheet['A' + str(row_termine)] = fnt_worksheet['A' + str(i)].value
        termine_worksheet['B' + str(row_termine)] = fnt_worksheet['E' + str(i)].value
        termine_worksheet['C' + str(row_termine)] = fnt_worksheet['G' + str(i)].value
        termine_worksheet['C' + str(row_termine)].number_format = 'DD.MM.YYYY'
        row_termine += 1
        if fnt_worksheet['E' + str(i)].value not in fnt_projects:
            fnt_projects.append(fnt_worksheet['E' + str(i)].value)
        if fnt_worksheet['E' + str(i)].value not in all_projects:
            all_projects.append(fnt_worksheet['E' + str(i)].value)
            check_worksheet['B' + str(row_check)] = fnt_worksheet['E' + str(i)].value
            row_check += 1
        
for i in range(2, ab_worksheet.max_row + 1):
    if not ab_worksheet['E' + str(i)].value:
        break
    if 'B2C' in ab_worksheet['E' + str(i)].value and ab_worksheet['G' + str(i)].value and 'Testprojekt4 B2C mit Vorlage' not in ab_worksheet['E' + str(i)].value:
        termine_worksheet['A' + str(row_termine)] = ab_worksheet['A' + str(i)].value
        termine_worksheet['B' + str(row_termine)] = ab_worksheet['E' + str(i)].value
        termine_worksheet['C' + str(row_termine)] = ab_worksheet['G' + str(i)].value
        termine_worksheet['C' + str(row_termine)].number_format = 'DD.MM.YYYY'
        row_termine += 1
        if ab_worksheet['E' + str(i)].value not in ab_projects:
            ab_projects.append(ab_worksheet['E' + str(i)].value)
        if ab_worksheet['E' + str(i)].value not in all_projects:
            all_projects.append(ab_worksheet['E' + str(i)].value)
            check_worksheet['B' + str(row_check)] = ab_worksheet['E' + str(i)].value
            row_check += 1

for i in range(2, ibn_worksheet.max_row + 1): 
    if not ibn_worksheet['E' + str(i)].value:
        break
    if 'B2C' in ibn_worksheet['E' + str(i)].value and ibn_worksheet['G' + str(i)].value and 'Testprojekt4 B2C mit Vorlage' not in ibn_worksheet['E' + str(i)].value:
        termine_worksheet['A' + str(row_termine)] = ibn_worksheet['A' + str(i)].value
        termine_worksheet['B' + str(row_termine)] = ibn_worksheet['E' + str(i)].value
        termine_worksheet['C' + str(row_termine)] = ibn_worksheet['G' + str(i)].value
        termine_worksheet['C' + str(row_termine)].number_format = 'DD.MM.YYYY'
        row_termine += 1
        if ibn_worksheet['E' + str(i)].value not in ibn_projects:
            ibn_projects.append(ibn_worksheet['E' + str(i)].value)
        if ibn_worksheet['E' + str(i)].value not in all_projects:
            all_projects.append(ibn_worksheet['E' + str(i)].value)
            check_worksheet['B' + str(row_check)] = ibn_worksheet['E' + str(i)].value
            row_check += 1

#insert a filter to the PoP Termine worksheet
termine_worksheet.auto_filter.ref = termine_worksheet.dimensions

#load the data into a DataFrame
df = pd.DataFrame(termine_worksheet.values)

#find the range of dates
oldest_date = df.iloc[2:].min()[2]
newest_date = df.iloc[2:].max()[2]
weeks_diff = math.ceil((newest_date - oldest_date).days / 7) + 1

#extract the header row
header_row = df.iloc[1]

#skip the header row and sort the data
sorted_df = df.iloc[2:].sort_values(by=[1, 2], ascending=[True, True])

#reset index after sorting
sorted_df.reset_index(drop=True, inplace=True)

#insert the header row back into the sorted DataFrame
sorted_df = pd.concat([header_row.to_frame().T, sorted_df], ignore_index=True)

#overwrite the worksheet with the sorted data, preserving styling and the header row
for r_idx, row in enumerate(dataframe_to_rows(sorted_df, index=False, header=False), 1):
    for c_idx, value in enumerate(row, 1):
        termine_worksheet.cell(row=r_idx + 1, column=c_idx, value=value)

#add calendar weeks, dates and quarters
week = oldest_date.isocalendar()[1]
year = oldest_date.year

# Loop through the columns
for i in range(4, 4 + weeks_diff):
    termine_worksheet[get_column_letter(i) + '2'] = week
    termine_worksheet[get_column_letter(i) + '2'].font = Font(b=True)
    termine_worksheet[get_column_letter(i) + '2'].alignment = Alignment(horizontal='center')

    for j in range(3, row_termine):
        # Extract date and calculate its week and year
        date_value = termine_worksheet['C' + str(j)].value
        date_week = date_value.isocalendar()[1]
        date_year = date_value.year
        
        # Ensure correct year and week alignment, especially at year-end
        if week == 1 and date_value.month == 12:
            date_year += 1
        elif week == 52 and date_value.month == 1:
            date_year -= 1

        # Check if the week and year match, then place the date in the correct cell
        if week == date_week and year == date_year:
            termine_worksheet[get_column_letter(i) + str(j)] = date_value
            termine_worksheet[get_column_letter(i) + str(j)].number_format = 'DD.MM.YYYY'
            termine_worksheet[get_column_letter(i) + str(j)].alignment = Alignment(horizontal='center')

    # Determine the quarter
    if week <= 13:
        quarter = 1
    elif week <= 26:
        quarter = 2
    elif week <= 39:
        quarter = 3
    else:
        quarter = 4

    # Set quarter and year in the first row
    termine_worksheet[get_column_letter(i) + '1'] = f'Q{quarter}/{year}'
    termine_worksheet[get_column_letter(i) + '1'].alignment = Alignment(horizontal='center')

    # Move to the next week, handling the transition to the next year
    week += 1
    if week > 52:
        week = 1
        year += 1

    # Adjust the column width
    termine_worksheet.column_dimensions[get_column_letter(i)].width = 12
    
#add KW    
for i in range(4, 4 + weeks_diff):
    termine_worksheet[get_column_letter(i) + '2'] = 'KW ' + str(termine_worksheet[get_column_letter(i) + '2'].value)

thick_top = Border(left=Side(style=None),
                     right=Side(style=None),
                     top=Side(style='medium'),
                     bottom=Side(style=None)
                     )

#add top border and color start and end dates of projects
project = None
boxes_start = []
boxes_end = []
for i in range(3, row_termine + 1):
    if project != termine_worksheet['B' + str(i)].value:
        for j in range(1, 4 + weeks_diff):
            termine_worksheet[get_column_letter(j) + str(i)].border = thick_top
            if 'KW' in termine_worksheet[get_column_letter(j) + '2'].value:
                if termine_worksheet[get_column_letter(j) + str(i)].value:
                    boxes_start.append((i, j))
            if 'KW' in termine_worksheet[get_column_letter(j) + '2'].value and i != 3:
                if termine_worksheet[get_column_letter(j) + str(i - 1)].value:
                    boxes_end.append((i - 1, j))
            
    project = termine_worksheet['B' + str(i)].value

# Iterate over the boxes_start and boxes_end lists
for start_coord, end_coord in zip(boxes_start, boxes_end):
    start_row, start_col = start_coord
    end_row, end_col = end_coord
    
    # Iterate over the range between start and end coordinates
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            termine_worksheet[get_column_letter(col) + str(row)].fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid') #light blue
            
    # Apply orange fill to the start and end cells
    termine_worksheet[get_column_letter(start_col) + str(start_row)].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid') #orange
    termine_worksheet[get_column_letter(end_col) + str(end_row)].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid') #orange
            
check_worksheet['A2'] = 'ID'
check_worksheet['B2'] = 'Projekt'
check_worksheet['C2'] = 'Eingang Doku'
check_worksheet['D2'] = 'Prüfung FNT'
check_worksheet['E2'] = 'Abnahme MFG'
check_worksheet['F2'] = 'IBN PoP'

#insert a filter to the PoP Check worksheet
check_worksheet.auto_filter.ref = check_worksheet.dimensions

#fill ID column
i = 1
while i < len(all_projects) + 1:
    check_worksheet['A' + str(i + 2)] = i
    check_worksheet['A' + str(i + 2)].alignment = Alignment(horizontal='center')
    i += 1

#fill Eingang Doku column
ed_missing = 0
for i in range(3, row_check):
    project_name = check_worksheet['B' + str(i)].value
    project_found = False
    for row in range(2, ed_worksheet.max_row + 1):
        if ed_worksheet['E' + str(row)].value == project_name:
            check_worksheet['C' + str(i)] = ed_worksheet['C' + str(row)].value
            project_found = True
            break
    
    if not project_found:
        ed_missing += 1
        check_worksheet['C' + str(i)].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid') #red

#fill Prüfung FNT column
fnt_missing = 0
for i in range(3, row_check):
    project_name = check_worksheet['B' + str(i)].value
    project_found = False
    for row in range(2, fnt_worksheet.max_row + 1):
        if fnt_worksheet['E' + str(row)].value == project_name:
            check_worksheet['D' + str(i)] = fnt_worksheet['C' + str(row)].value
            project_found = True
            break
    
    if not project_found:
        fnt_missing += 1
        check_worksheet['D' + str(i)].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid') #red

#fill Abnahme MFG column
ab_missing = 0
for i in range(3, row_check):
    project_name = check_worksheet['B' + str(i)].value
    project_found = False
    for row in range(2, ab_worksheet.max_row + 1):
        if ab_worksheet['E' + str(row)].value == project_name:
            check_worksheet['E' + str(i)] = ab_worksheet['C' + str(row)].value
            project_found = True
            break
    
    if not project_found:
        ab_missing += 1
        check_worksheet['E' + str(i)].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid') #red

#fill IBN PoP column 
ibn_missing = 0       
for i in range(3, row_check):
    project_name = check_worksheet['B' + str(i)].value
    project_found = False
    for row in range(2, ibn_worksheet.max_row + 1):
        if ibn_worksheet['E' + str(row)].value == project_name:
            check_worksheet['F' + str(i)] = ibn_worksheet['C' + str(row)].value
            project_found = True
            break
    
    if not project_found:
        ibn_missing += 1
        check_worksheet['F' + str(i)].fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid') #red

#input the count of projects
check_worksheet['B1'] = len(all_projects)
check_worksheet['C1'] = ed_missing
check_worksheet['D1'] = fnt_missing
check_worksheet['E1'] = ab_missing
check_worksheet['F1'] = ibn_missing

#change style of headers and counts
for i in range(1, 7):
    check_worksheet.column_dimensions[get_column_letter(i)].width = 20
    check_worksheet[get_column_letter(i) + '2'].font = Font(b=True)
    check_worksheet[get_column_letter(i) + '1'].font = Font(b=True)
    check_worksheet[get_column_letter(i) + '1'].alignment = Alignment(horizontal='center')
    termine_worksheet[get_column_letter(i) + '2'].font = Font(b=True)
    termine_worksheet[get_column_letter(i) + '2'].alignment = Alignment(horizontal='center')

check_worksheet.column_dimensions['B'].width = 62
check_worksheet['A2'].alignment = Alignment(horizontal='center')

termine_worksheet.column_dimensions['B'].width = 62
termine_worksheet.column_dimensions['C'].width = 12

# Define the font you want to set as default
default_font_name = 'Arial'
default_font_size = 11

# Iterate over each worksheet in the workbook
for ws in ot_file.worksheets:
    # Apply the font to all cells in the worksheet
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            existing_font = cell.font
            # Create a new Font object preserving the existing attributes
            new_font = Font(name=default_font_name, 
                            size=default_font_size,
                            bold=existing_font.bold,
                            italic=existing_font.italic,
                            vertAlign=existing_font.vertAlign,
                            underline=existing_font.underline,
                            strike=existing_font.strike,
                            color=existing_font.color)
            cell.font = new_font

ot_file_export = os.path.join(directory, os.path.basename(ot_file_path).split('.')[0] + '_v2.xlsx')

ot_file.save(ot_file_export)

print(f"OT File has been exported to:\n'{ot_file_export}'.")
changelog += f"OT File has been exported to:\n '{ot_file_export}'.\n"

# Calculate the elapsed time
elapsed_time = time.time() - start_time

# Convert elapsed time to minutes and seconds
minutes = int(elapsed_time // 60)
seconds = int(elapsed_time % 60)

def show_popup():
    if minutes > 0:
        messagebox.showinfo("Script Completed", changelog + f'\nElapsed Time: {minutes} min, {seconds} s.')
    else:
        messagebox.showinfo("Script Completed", changelog + f'\nElapsed Time: {seconds} s.')
    
# Call this function after the script has completed running
show_popup()